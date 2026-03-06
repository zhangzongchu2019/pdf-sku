"""
Excel 导出器 — 将 Job 的 SKU 识别结果导出为两个 Excel 文件。

File 1 (full_export.xlsx):   商品子图 + 所有提取到的 SKU 属性（动态列）
File 2 (keywords_export.xlsx): 固定关键词模板（平台导入格式）

多图支持 (keywords_export): 同一 SKU 的多张商品子图各占一列，
  列名均为"商品图片"（不编号，可重复）。
"""
from __future__ import annotations

import io
import json
from dataclasses import dataclass, field
from pathlib import Path
from uuid import UUID

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

import structlog

logger = structlog.get_logger()

# ─────────────────────── 关键词列定义 ───────────────────────
# 平台导入模板列顺序（不含首列"商品图片"）
# 每项: (列标题, 语义描述)  语义描述为空字符串的列固定留空
KEYWORD_FIELDS: list[tuple[str, str]] = [
    ('规格图片',           ''),
    ('商品名称/描述',       '商品型号 + 规格拼接，例如"WS X-683 950W*760D*850H"'),
    ('售价',              '商品零售单价或市场价，面向终端消费者'),
    ('货号',              '商品型号/货号编号，例如"A105#"、"WS X-683"'),
    ('商品ID',            ''),
    ('标签',              ''),
    ('来源(仅自己可见)',    ''),    # 特殊列: PDF 文件名
    ('商品简称',           ''),
    ('商品规格',           '商品的规格、尺寸，例如"130x70x55cm"'),
    ('颜色',              '商品颜色或色号，例如"红色"、"深蓝"'),
    ('规格编码',           ''),
    ('批发价',             '批量采购价格，通常低于零售价'),
    ('打包价',             '打包装或组合销售价格'),
    ('代发价',             '供货商代发货价格（dropship），直接发给终端买家'),
    ('拿货价(仅自己可见)', ''),
    ('活动类型',           ''),
    ('活动价',             '促销、活动或限时优惠价格'),
    ('库存',              '商品库存数量'),
    ('重量(kg)',           '商品重量，单位 kg'),
    ('备注(公开)',         ''),
    ('自动下架时间',        '商品自动下架或到期时间'),
]

# 固定留空的列
_EMPTY_KW_COLS: frozenset[str] = frozenset({
    '规格图片', '商品ID', '标签', '商品简称',
    '规格编码', '拿货价(仅自己可见)', '活动类型', '备注(公开)',
})

# 特殊处理列（不走 LLM 映射，从 ExportRow 字段直接计算）
_SOURCE_COL = '来源(仅自己可见)'   # PDF 文件名
_NAME_COL   = '商品名称/描述'      # 型号 + 规格
_MODEL_COL  = '货号'               # 型号

# 固定列的候选键集合（用于从动态列中去除已覆盖的字段）
_FIXED_CANDIDATES: set[str] = {
    'product_name', 'name', '产品名称',
    'spec', 'specification', 'size', '规格', '尺寸',
    'unit_price', 'price', '单价',
    'color', 'colour', '颜色',
    'remark', 'note', '备注', '说明',
    'weight', 'weight_kg', '重量',
}


# ─────────────────────── 数据结构 ───────────────────────

@dataclass
class ExportRow:
    page_number: int
    sku_id: str
    attributes: dict
    images: list[bytes] = field(default_factory=list)   # 多图：每张占一列
    image_ids: list[str] = field(default_factory=list)
    source_text: str = ""   # PDF 原始 OCR 文字（按页）
    variant_label: str = ""  # 座位/规格变体标签，如"1人位"/"2人位"


# ─────────────────────── 辅助函数 ───────────────────────

def _get_field_value(attrs: dict, candidates: list[str]) -> str:
    """遍历候选键列表，返回第一个命中的非空值，否则返回空字符串。"""
    for key in candidates:
        val = attrs.get(key)
        if val is not None and val != "":
            return str(val)
    return ""


IMG_MAX_PX = 400          # 导出图片最大边长（像素）
IMG_COL_WIDTH = 57        # 图片列宽（Excel 字符单位，约 400px）
IMG_ROW_HEIGHT = 300      # 图片行高（Excel point 单位，约 400px）


def _preprocess_image(img_bytes: bytes, max_px: int = IMG_MAX_PX) -> bytes | None:
    """PIL 缩放图片为 PNG bytes，线程安全，可并行执行。"""
    if not img_bytes:
        return None
    try:
        from PIL import Image as PILImage
        pil = PILImage.open(io.BytesIO(img_bytes))
        if pil.width > max_px or pil.height > max_px:
            pil.thumbnail((max_px, max_px), PILImage.LANCZOS)
        buf = io.BytesIO()
        pil.save(buf, format="PNG")
        return buf.getvalue()
    except Exception as e:
        logger.warning("excel_preprocess_image_failed", error=str(e))
        return None


def _embed_png(ws, png_bytes: bytes, row: int, col: int) -> None:
    """将已预处理的 PNG bytes 嵌入单元格（非线程安全，须在主线程顺序调用）。"""
    try:
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter
        xl_img = XLImage(io.BytesIO(png_bytes))
        xl_img.anchor = f"{get_column_letter(col)}{row}"
        ws.add_image(xl_img)
    except Exception as e:
        logger.warning("excel_embed_image_failed", row=row, col=col, error=str(e))


def _embed_image(ws, img_bytes: bytes, row: int, col: int = 1, max_px: int = IMG_MAX_PX) -> None:
    """将图片 bytes 缩放后嵌入到指定单元格（保留向后兼容）。"""
    png = _preprocess_image(img_bytes, max_px)
    if png:
        _embed_png(ws, png, row, col)


def _preprocess_rows_images(
    rows: list, n_img_cols: int, max_px: int = IMG_MAX_PX
) -> dict[tuple[int, int], bytes]:
    """并行预处理所有行的图片，返回 {(row_idx, col): png_bytes} 字典。"""
    from concurrent.futures import ThreadPoolExecutor

    tasks: list[tuple[int, int, bytes]] = []
    for row_idx, row in enumerate(rows, 2):
        for img_idx, img_bytes in enumerate(row.images[:n_img_cols]):
            if img_bytes:
                tasks.append((row_idx, img_idx + 1, img_bytes))

    if not tasks:
        return {}

    with ThreadPoolExecutor(max_workers=min(8, len(tasks))) as pool:
        processed = list(pool.map(lambda t: _preprocess_image(t[2], max_px), tasks))

    return {
        (ri, ci): png
        for (ri, ci, _), png in zip(tasks, processed)
        if png
    }


def _apply_header_style(ws, headers: list[str], n_img_cols: int = 1) -> None:
    """统一设置首行表头样式：浅蓝底色 + 加粗 + 居中。图片列宽 14，文字列宽 20。"""
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="BDD7EE")
    header_font = Font(bold=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 图片列宽 IMG_COL_WIDTH，其余列宽 20
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        if col_idx <= n_img_cols:
            ws.column_dimensions[col_letter].width = IMG_COL_WIDTH
        else:
            ws.column_dimensions[col_letter].width = 20

    ws.freeze_panes = f"{get_column_letter(n_img_cols + 1)}2"


# ─────────────────────── LLM 语义映射 ───────────────────────

async def build_keyword_mapping_via_llm(
    rows: list[ExportRow],
    llm_service,
) -> dict[str, list[str]]:
    """
    通过一次 LLM 调用，将 SKU 属性键语义映射到标准关键词列。

    返回: {关键词列标题: [匹配的属性键, ...]}（按置信度从高到低）
    如某列无匹配，则对应值为空列表。
    """
    # 1. 收集所有属性键及各自的示例值（最多 3 个不同值）
    key_samples: dict[str, list[str]] = {}
    for row in rows:
        for k, v in row.attributes.items():
            if v is None or v == "":
                continue
            samples = key_samples.setdefault(k, [])
            sv = str(v).strip()
            if sv and sv not in samples and len(samples) < 3:
                samples.append(sv)

    if not key_samples:
        return {kf[0]: [] for kf in KEYWORD_FIELDS}

    # 2. 构造 LLM prompt（仅映射有语义描述的非空列）
    mappable_fields = [
        kf for kf in KEYWORD_FIELDS
        if kf[1]  # 语义描述非空 → 需要 LLM 映射
        and kf[0] not in (_NAME_COL, _MODEL_COL)  # 这两列直接从属性提取，不走 LLM
    ]

    attr_lines = "\n".join(
        f'  "{k}": {json.dumps(v, ensure_ascii=False)}'
        for k, v in key_samples.items()
    )
    keyword_lines = "\n".join(
        f'  "{kf[0]}": {kf[1]}'
        for kf in mappable_fields
    )

    prompt = f"""你是一个商品数据结构专家。请根据语义理解，将 SKU 属性字段映射到标准导出列。

## SKU 属性字段（字段名: [示例值列表]）
{{{attr_lines}
}}

## 标准导出列（列名: 语义说明）
{{{keyword_lines}
}}

## 任务
对于每个标准导出列，从属性字段中找出语义上最匹配的字段名（可能有多个，按匹配度从高到低排列）。
如果没有任何字段与某个列语义相关，则返回空数组 []。

## 返回格式（严格 JSON，不要有任何额外文字）
{{
  "售价": ["最匹配的字段名", "次匹配的字段名"],
  ...（覆盖所有 {len(mappable_fields)} 个标准列）
}}"""

    try:
        resp = await llm_service.call_llm(
            operation="keyword_mapping",
            prompt=prompt,
        )

        # 3. 解析 LLM 返回的 JSON
        raw = resp.content.strip()
        # 兼容 LLM 可能输出 ```json ... ``` 包裹
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        mapping: dict = json.loads(raw.strip())

        # 4. 验证并规范化：仅保留 mappable_fields 中存在的列，且值必须是列表
        result: dict[str, list[str]] = {}
        for col_name, _ in KEYWORD_FIELDS:
            matched = mapping.get(col_name, [])
            if not isinstance(matched, list):
                matched = [str(matched)] if matched else []
            # 只保留实际存在于数据中的属性键
            matched = [k for k in matched if k in key_samples]
            result[col_name] = matched

        logger.info(
            "keyword_mapping_llm_done",
            attr_keys=len(key_samples),
            mapped_cols=sum(1 for v in result.values() if v),
        )
        return result

    except Exception as e:
        logger.warning("keyword_mapping_llm_failed", error=str(e))
        return {kf[0]: [] for kf in KEYWORD_FIELDS}


def _apply_keyword_mapping(attrs: dict, keyword_mapping: dict[str, list[str]]) -> dict[str, str]:
    """
    给定单行 SKU 的 attributes 和 LLM 生成的 keyword_mapping，
    提取每个关键词列对应的值。
    """
    result: dict[str, str] = {}
    for col_name, _ in KEYWORD_FIELDS:
        candidates = keyword_mapping.get(col_name, [])
        value = ""
        for key in candidates:
            v = attrs.get(key)
            if v is not None and str(v).strip():
                value = str(v).strip()
                break
        result[col_name] = value
    return result


# ─────────────────────── LLM 映射缓存 ───────────────────────
# key: job_id 字符串，value: keyword_mapping dict
_keyword_mapping_cache: dict[str, dict[str, list[str]]] = {}


# ─────────────────────── 主类 ───────────────────────

class ExcelExporter:
    def __init__(self, job_data_dir: str) -> None:
        self._job_data_dir = Path(job_data_dir)

    async def load_job_data(self, db: AsyncSession, job_id: UUID) -> list[ExportRow]:
        """
        查询 SKU + 绑定图片，构建导出行列表。

        合并规则：
        - 相同 product_id 的 SKU → 合并为一行（同一商品的不同规格）
        - 无 product_id 但绑定相同首图的 SKU → 合并为一行
        - 仅输出完整商品子图（is_fragmented=False, search_eligible=True）
        - 不输出未绑定 SKU 的游离图片行
        """
        from pdf_sku.common.models import SKU, SKUImageBinding, Image, PDFJob
        from collections import defaultdict

        # 1. 所有非 superseded 的 SKU，按页码 + 插入顺序排序
        skus = (await db.execute(
            select(SKU)
            .where(SKU.job_id == job_id, SKU.superseded == False)
            .order_by(SKU.page_number, SKU.id)
        )).scalars().all()

        sku_ids = [s.sku_id for s in skus]

        # 2. 每个 SKU 的有效图片绑定，按置信度降序
        sku_image_ids: dict[str, list[str]] = defaultdict(list)
        if sku_ids:
            bindings = (await db.execute(
                select(SKUImageBinding)
                .where(
                    SKUImageBinding.job_id == job_id,
                    SKUImageBinding.sku_id.in_(sku_ids),
                    SKUImageBinding.is_latest == True,
                    SKUImageBinding.image_id.isnot(None),
                )
                .order_by(
                    SKUImageBinding.sku_id,
                    SKUImageBinding.binding_confidence.desc(),
                    SKUImageBinding.image_id,
                )
            )).scalars().all()

            for b in bindings:
                if b.image_id and b.image_id not in sku_image_ids[b.sku_id]:
                    sku_image_ids[b.sku_id].append(b.image_id)

        # 3. 批量查询图片元数据
        # all_images: 用于分组判断（所有绑定的图片，不过滤）
        # display_images: 用于实际展示（仅完整商品子图，过滤瓦片碎片）
        all_bound_image_ids = list({iid for ids in sku_image_ids.values() for iid in ids})
        all_images: dict[str, object] = {}
        display_images: set[str] = set()
        if all_bound_image_ids:
            img_rows = (await db.execute(
                select(Image)
                .where(Image.job_id == job_id, Image.image_id.in_(all_bound_image_ids))
            )).scalars().all()
            for img in img_rows:
                all_images[img.image_id] = img
                # 仅完整商品子图参与展示（排除瓦片碎片）
                if not img.is_fragmented and img.search_eligible:
                    display_images.add(img.image_id)

        job_dir = self._job_data_dir / str(job_id)

        # 4. 查询 PDF 原始文件名
        source_filename = ""
        try:
            pdf_job = (await db.execute(
                select(PDFJob).where(PDFJob.job_id == job_id)
            )).scalar_one_or_none()
            if pdf_job:
                source_filename = pdf_job.source_file or ""
        except Exception as e:
            logger.warning("excel_source_filename_failed", error=str(e))

        def _load_image(image_id: str) -> bytes | None:
            img = all_images.get(image_id)
            if not img:
                return None
            try:
                return (job_dir / img.extracted_path).read_bytes()
            except Exception as e:
                logger.warning("excel_export_image_read_failed",
                               image_id=image_id, error=str(e))
                return None

        def _build_row(sku_group: list, group_id: str) -> ExportRow:
            """将一组 SKU 合并为一个 ExportRow。展示图仅含完整子图。"""
            # 按 SKU 顺序收集去重后的图片，仅展示完整子图（非碎片）
            ordered_ids: list[str] = []
            seen: set[str] = set()
            for sku in sku_group:
                for iid in sku_image_ids.get(sku.sku_id, []):
                    if iid not in seen and iid in display_images:
                        seen.add(iid)
                        ordered_ids.append(iid)

            image_bytes_list: list[bytes] = []
            for iid in ordered_ids:
                data = _load_image(iid)
                if data:
                    image_bytes_list.append(data)

            # 代表 SKU（第一个）的属性作为基准
            rep = sku_group[0]
            attrs = dict(rep.attributes or {})

            # 规格拼接：variant_label + size
            variant_parts: list[str] = []
            for sku in sku_group:
                vl = sku.variant_label or ""
                a = sku.attributes or {}
                size = str(
                    a.get("size") or a.get("规格") or
                    a.get("spec") or a.get("specification") or ""
                ).strip()
                if vl and size:
                    variant_parts.append(f"{vl}: {size}")
                elif vl:
                    variant_parts.append(vl)
                elif size:
                    variant_parts.append(size)

            variant_label = (
                " / ".join(variant_parts)
                if len(sku_group) > 1 and variant_parts
                else (rep.variant_label or "")
            )

            return ExportRow(
                page_number=rep.page_number,
                sku_id=group_id,
                attributes=attrs,
                images=image_bytes_list,
                image_ids=ordered_ids,
                source_text=source_filename,
                variant_label=variant_label,
            )

        # 5. 对所有 SKU 统一用 Union-Find 分组
        #    规则1: 共享任意同一张图片的 SKU → 合并（相同商品子图 → 同一行）
        #    规则2: 同一 product_id 的 SKU → 强制合并（同商品不同规格 → 同一行）
        parent: dict[str, str] = {sku.sku_id: sku.sku_id for sku in skus}

        def _find(x: str) -> str:
            root = x
            while parent.get(root, root) != root:
                root = parent[root]
            # 路径压缩
            while parent.get(x, x) != root:
                parent[x], x = root, parent[x]
            return root

        def _union(x: str, y: str) -> None:
            px, py = _find(x), _find(y)
            if px != py:
                parent[px] = py

        # 规则1: 共享图片合并（用 all_images 确保所有绑定图都参与）
        img_to_sku_ids: dict[str, list[str]] = defaultdict(list)
        for sku in skus:
            for iid in sku_image_ids.get(sku.sku_id, []):
                if iid in all_images:
                    img_to_sku_ids[iid].append(sku.sku_id)

        for sid_list in img_to_sku_ids.values():
            for i in range(1, len(sid_list)):
                _union(sid_list[0], sid_list[i])

        # 规则2: 同 product_id 强制合并
        pid_first: dict[str, str] = {}
        for sku in skus:
            if sku.product_id:
                if sku.product_id in pid_first:
                    _union(pid_first[sku.product_id], sku.sku_id)
                else:
                    pid_first[sku.product_id] = sku.sku_id

        # 按 root 聚合（保持原始 SKU 顺序）
        union_groups: dict[str, list] = defaultdict(list)
        for sku in skus:
            union_groups[_find(sku.sku_id)].append(sku)

        rows: list[ExportRow] = []
        for grp in union_groups.values():
            # group_id: 优先用组内第一个 product_id，否则用第一个 sku_id
            group_id = next((s.product_id for s in grp if s.product_id), grp[0].sku_id)
            rows.append(_build_row(grp, group_id))

        rows.sort(key=lambda r: r.page_number)

        logger.info("excel_export_rows_loaded",
                    job_id=str(job_id),
                    total_rows=len(rows),
                    total_skus=len(skus),
                    union_groups=len(union_groups))
        return rows

    @staticmethod
    def build_full_excel_sync(rows: list[ExportRow]) -> io.BytesIO:
        """
        File 1: 商品图片1 | 商品图片2 | ... | 页码 | SKU ID | 固定属性列 | 动态属性列
        多图支持: 每张子图占一列，最大列数 = max(len(row.images))。
        """
        import openpyxl
        from openpyxl.styles import Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "全量导出"

        # 确定最多图片数
        n_img_cols = max((len(row.images) for row in rows), default=1)
        n_img_cols = max(n_img_cols, 1)

        # 图片列表头
        if n_img_cols == 1:
            img_headers = ["商品图片"]
        else:
            img_headers = [f"商品图片{i + 1}" for i in range(n_img_cols)]

        # 收集所有属性键，按频率降序
        key_freq: dict[str, int] = {}
        for row in rows:
            for k in row.attributes:
                key_freq[k] = key_freq.get(k, 0) + 1
        all_keys_by_freq = sorted(key_freq, key=lambda k: -key_freq[k])

        # 固定列定义
        fixed_cols: list[tuple[str, list[str]]] = [
            ("产品名称", ["product_name", "name", "产品名称"]),
            ("规格",     ["spec", "specification", "size", "规格", "尺寸"]),
            ("单价",     ["unit_price", "price", "单价"]),
            ("颜色",     ["color", "colour", "颜色"]),
            ("备注",     ["remark", "note", "备注", "说明"]),
            ("重量",     ["weight", "weight_kg", "重量"]),
        ]
        # 判断是否有 variant_label 数据，有则加入固定列
        has_variant = any(row.variant_label for row in rows)

        # 动态列：排除已被固定列覆盖的键
        extra_keys = [k for k in all_keys_by_freq if k not in _FIXED_CANDIDATES]

        variant_col = ["变体规格"] if has_variant else []
        headers = (
            img_headers
            + ["页码", "SKU ID"]
            + variant_col
            + [col[0] for col in fixed_cols]
            + extra_keys
        )
        _apply_header_style(ws, headers, n_img_cols=n_img_cols)

        # 并行预处理所有图片
        preprocessed = _preprocess_rows_images(rows, n_img_cols)

        # 数据行
        for row_idx, row in enumerate(rows, 2):
            attrs = row.attributes

            # 文字列（图片列之后）
            text_data: list = [
                row.page_number,
                row.sku_id,
            ]
            if has_variant:
                text_data.append(row.variant_label)
            for _, candidates in fixed_cols:
                text_data.append(_get_field_value(attrs, candidates))
            for k in extra_keys:
                v = attrs.get(k)
                text_data.append(str(v) if v is not None else "")

            # 写文字列（从第 n_img_cols+1 列开始）
            for col_offset, val in enumerate(text_data):
                ws.cell(
                    row=row_idx,
                    column=n_img_cols + 1 + col_offset,
                    value=val,
                ).alignment = Alignment(vertical="center", wrap_text=True)

            # 嵌入已预处理图片（每张占一列）
            has_image = False
            for img_idx in range(min(len(row.images), n_img_cols)):
                png = preprocessed.get((row_idx, img_idx + 1))
                if png:
                    _embed_png(ws, png, row_idx, img_idx + 1)
                    has_image = True

            ws.row_dimensions[row_idx].height = IMG_ROW_HEIGHT if has_image else 15

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def build_full_excel(self, rows: list[ExportRow]) -> io.BytesIO:
        return self.build_full_excel_sync(rows)

    @staticmethod
    def build_keywords_excel_sync(
        rows: list[ExportRow],
        keyword_mapping: dict | None,
    ) -> io.BytesIO:
        """纯同步构建关键词 Excel，供线程池调用。keyword_mapping 已由外部预先获取。"""
        import openpyxl
        from openpyxl.styles import Alignment

        n_img_cols = max((len(row.images) for row in rows), default=1)
        n_img_cols = max(n_img_cols, 1)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "关键词导出"

        headers = ["商品图片"] * n_img_cols + [kf[0] for kf in KEYWORD_FIELDS]
        _apply_header_style(ws, headers, n_img_cols=n_img_cols)

        _FALLBACK_CANDIDATES: dict[str, list[str]] = {
            '售价':          ['unit_price', 'price', 'retail_price', '单价', '售价'],
            '商品规格':      ['size', 'spec', 'specification', '规格', '尺寸'],
            '颜色':          ['color', 'colour', '颜色', '色号'],
            '批发价':        ['wholesale_price', '批发价', '批发'],
            '打包价':        ['package_price', '打包价', '打包'],
            '代发价':        ['dropship_price', '代发价', '代发'],
            '活动价':        ['promotion_price', 'activity_price', '活动价', '促销价'],
            '库存':          ['stock', 'inventory', '库存', '数量'],
            '重量(kg)':      ['weight', '重量', 'weight_kg', '毛重'],
            '自动下架时间':  ['auto_offline_time', '下架时间', 'offline_time'],
        }

        # 并行预处理所有图片
        preprocessed = _preprocess_rows_images(rows, n_img_cols)

        for row_idx, row in enumerate(rows, 2):
            attrs = row.attributes
            model = _get_field_value(attrs, ['model_number', 'model', '型号'])
            pname = _get_field_value(attrs, ['product_name', '产品名称', '品名', 'name'])
            size  = _get_field_value(attrs, ['size', 'spec', 'specification', '规格', '尺寸'])
            variant = row.variant_label
            name_parts = []
            if pname and pname != model:
                name_parts.append(pname)
            name_parts.append(model)
            if variant:
                name_parts.append(variant)
            name_parts.append(size)
            product_name_val = " ".join(p for p in name_parts if p)

            kw_data: list[str] = []
            for col_name, _ in KEYWORD_FIELDS:
                if col_name in _EMPTY_KW_COLS:
                    kw_data.append("")
                elif col_name == _SOURCE_COL:
                    kw_data.append(row.source_text)
                elif col_name == _NAME_COL:
                    kw_data.append(product_name_val)
                elif col_name == _MODEL_COL:
                    kw_data.append(model)
                elif keyword_mapping is not None:
                    candidates = keyword_mapping.get(col_name, [])
                    value = ""
                    for key in candidates:
                        v = attrs.get(key)
                        if v is not None and str(v).strip():
                            value = str(v).strip()
                            break
                    kw_data.append(value)
                else:
                    candidates = _FALLBACK_CANDIDATES.get(col_name, [])
                    kw_data.append(_get_field_value(attrs, candidates))

            for col_offset, val in enumerate(kw_data):
                ws.cell(
                    row=row_idx,
                    column=n_img_cols + 1 + col_offset,
                    value=val,
                ).alignment = Alignment(vertical="center", wrap_text=True)

            has_image = False
            for img_idx in range(min(len(row.images), n_img_cols)):
                png = preprocessed.get((row_idx, img_idx + 1))
                if png:
                    _embed_png(ws, png, row_idx, img_idx + 1)
                    has_image = True

            ws.row_dimensions[row_idx].height = IMG_ROW_HEIGHT if has_image else 15

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    async def build_keywords_excel(
        self,
        rows: list[ExportRow],
        llm_service=None,
        cache_key: str | None = None,
    ) -> io.BytesIO:
        """关键词导出 Excel（平台导入格式）。LLM 映射支持 cache_key 缓存，构建在线程池执行。"""
        import asyncio

        keyword_mapping: dict[str, list[str]] | None = None
        if llm_service is not None and rows:
            if cache_key and cache_key in _keyword_mapping_cache:
                keyword_mapping = _keyword_mapping_cache[cache_key]
                logger.info("keywords_excel_using_cached_mapping", cache_key=cache_key)
            else:
                keyword_mapping = await build_keyword_mapping_via_llm(rows, llm_service)
                if cache_key:
                    _keyword_mapping_cache[cache_key] = keyword_mapping
                logger.info("keywords_excel_using_llm_mapping")
        else:
            logger.info("keywords_excel_using_fallback_mapping")

        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(
            None,
            self.build_keywords_excel_sync,
            rows,
            keyword_mapping,
        )
